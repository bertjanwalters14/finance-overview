"""
Eenmalig hulpscript: leest seed-data/Financiën.xlsx en zet de historische
gegevens om naar seed-data/seed.json, in het datamodel dat de app gebruikt
(zie src/lib/types.ts). Run: python seed-data/parse_excel_seed.py
"""

import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, "Financiën.xlsx")
OUT_PATH = os.path.join(HERE, "seed.json")

MAAND_NUMMERS = {
    "januari": 1,
    "februari": 2,
    "maart": 3,
    "april": 4,
    "mei": 5,
    "juni": 6,
    "juli": 7,
    "augustus": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "december": 12,
}


def maand_nummer(naam):
    return MAAND_NUMMERS[str(naam).strip().lower()]


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

seed = {
    "maanden": [],
    "vermogen": None,
    "aandelenPayt": [],
    "loonontwikkeling": [],
    "doelen": [],
}

# --- Financiën sheet: vaste lasten template, huidig-jaar doelen, vermogen ---
fin = wb["Financiën"]

vaste_lasten = []
loon = 0.0
for row in range(2, 9):
    naam = fin.cell(row=row, column=3).value
    bedrag = fin.cell(row=row, column=4).value
    if naam is not None and bedrag is not None:
        vaste_lasten.append({"naam": naam, "bedrag": float(bedrag)})
    if fin.cell(row=row, column=1).value == "Loon":
        loon = float(fin.cell(row=row, column=2).value)

# Maand/Doel/Werkelijk/Belegging tabel in kolom H-K, huidig jaar (2026)
HUIDIG_JAAR = 2026
doel_per_maand = [0.0] * 12
werkelijk_per_maand = [0.0] * 12
belegging_per_maand = [0.0] * 12
maanden_2026 = []

for row in range(2, 14):
    maand_naam = fin.cell(row=row, column=8).value
    if not maand_naam:
        continue
    m = maand_nummer(maand_naam)
    doel = float(fin.cell(row=row, column=9).value or 0)
    werkelijk = float(fin.cell(row=row, column=10).value or 0)
    belegging = float(fin.cell(row=row, column=11).value or 0)
    doel_per_maand[m - 1] = doel
    werkelijk_per_maand[m - 1] = werkelijk
    belegging_per_maand[m - 1] = belegging
    # Alle 12 maanden van het huidige jaar worden als Maand-record aangemaakt
    # (net als in de Excel), ook toekomstige maanden met alleen een doel en
    # nog geen werkelijke cijfers.
    maanden_2026.append(
        {
            "jaar": HUIDIG_JAAR,
            "maand": m,
            "loon": loon,
            "overigeInkomsten": 0.0,
            "vasteLasten": vaste_lasten,
            "doelSparen": doel,
            "werkelijkGespaard": werkelijk,
            "beleggingInleg": belegging,
        }
    )

seed["maanden"] = maanden_2026

seed["doelen"].append(
    {
        "jaar": HUIDIG_JAAR,
        "doelPerMaand": doel_per_maand,
        "werkelijkPerMaand": werkelijk_per_maand,
        "categorieen": [
            {"naam": "Belegging", "bedragenPerMaand": belegging_per_maand}
        ],
    }
)

# Vermogen snapshot
def cell(sheet, row, col):
    return sheet.cell(row=row, column=col).value

spaarrekening = float(cell(fin, 21, 2))
belegging_totaal = float(cell(fin, 22, 2))
aandelen_payt_waarde = float(cell(fin, 26, 3))  # 'Bedrag' kolom van eigen aandelen
huis_waarde = float(cell(fin, 27, 2))
hypotheek = float(cell(fin, 28, 2))
overwaarde_aandeel = float(cell(fin, 29, 3))
schuld = float(cell(fin, 33, 2))

seed["vermogen"] = {
    "spaarrekening": spaarrekening,
    "belegging": belegging_totaal,
    "aandelenPaytWaarde": aandelen_payt_waarde,
    "huisWaarde": huis_waarde,
    "hypotheek": hypotheek,
    "overwaardeAandeel": overwaarde_aandeel,
    "schuld": schuld,
    "bijgewerktOp": "2026-08-05",
}

# --- Aandelen Payt sheet ---
aandelen = wb["Aandelen Payt"]
for row in range(2, 6):
    naam = cell(aandelen, row, 1)
    if not naam:
        continue
    seed["aandelenPayt"].append(
        {
            "naam": naam,
            "aantal": float(cell(aandelen, row, 2) or 0),
            "inleg": float(cell(aandelen, row, 5) or 0),
            "waarde": float(cell(aandelen, row, 6) or 0),
            "rendement": float(cell(aandelen, row, 9) or 0),
            "dividend": float(cell(aandelen, row, 10) or 0),
        }
    )

# --- Loonontwikkeling sheet ---
loon_sheet = wb["loon ontwikkeling"]
for row in range(1, 15):
    werkgever = cell(loon_sheet, row, 1)
    bedrag = cell(loon_sheet, row, 2)
    jaar = cell(loon_sheet, row, 4)
    if werkgever is None or bedrag is None or jaar is None:
        continue
    seed["loonontwikkeling"].append(
        {
            "jaar": int(jaar),
            "werkgever": str(werkgever).strip(),
            "bedrag": float(bedrag),
        }
    )

# --- Jaarlijkse doelen-sheets ---
DOELEN_SHEETS = {
    "Doelen 2025": {
        "jaar": 2025,
        "maand_col": 1,
        "doel_col": 2,
        "werkelijk_col": 3,
        "categorieen": [("Belegging", 4), ("Vakantie", 5)],
        "start_row": 2,
    },
    "2022 doelen": {
        "jaar": 2022,
        "maand_col": 1,
        "doel_col": 2,
        "werkelijk_col": 3,
        "categorieen": [("Belegging", 4), ("Vakantie", 5)],
        "start_row": 2,
    },
    "2021 doelen": {
        "jaar": 2021,
        "maand_col": 1,
        "doel_col": 2,
        "werkelijk_col": 3,
        "categorieen": [("Belegging", 5)],
        "start_row": 2,
    },
    "2020 doelen": {
        "jaar": 2020,
        "maand_col": 2,
        "doel_col": 3,
        "werkelijk_col": 4,
        "categorieen": [("Belegging", 6)],
        "start_row": 2,
    },
    # Doelen 2024 heeft geen headerrij; kolomindeling afgeleid uit het patroon
    # van de andere jaren (D=maand, E=doel, F=werkelijk, G=Belegging, H=Vakantie).
    # Controleer deze labels zelf nog even in de app.
    "Doelen 2024": {
        "jaar": 2024,
        "maand_col": 4,
        "doel_col": 5,
        "werkelijk_col": 6,
        "categorieen": [("Belegging", 7), ("Vakantie", 8)],
        "start_row": 1,
    },
}

for sheet_name, cfg in DOELEN_SHEETS.items():
    sheet = wb[sheet_name]
    doel_per_maand = [0.0] * 12
    werkelijk_per_maand = [0.0] * 12
    cat_data = {naam: [0.0] * 12 for naam, _ in cfg["categorieen"]}

    for row in range(cfg["start_row"], cfg["start_row"] + 12):
        maand_naam = cell(sheet, row, cfg["maand_col"])
        if not maand_naam:
            continue
        m = maand_nummer(maand_naam)
        doel_per_maand[m - 1] = float(cell(sheet, row, cfg["doel_col"]) or 0)
        werkelijk_per_maand[m - 1] = float(
            cell(sheet, row, cfg["werkelijk_col"]) or 0
        )
        for naam, col in cfg["categorieen"]:
            cat_data[naam][m - 1] = float(cell(sheet, row, col) or 0)

    seed["doelen"].append(
        {
            "jaar": cfg["jaar"],
            "doelPerMaand": doel_per_maand,
            "werkelijkPerMaand": werkelijk_per_maand,
            "categorieen": [
                {"naam": naam, "bedragenPerMaand": cat_data[naam]}
                for naam, _ in cfg["categorieen"]
            ],
        }
    )

seed["doelen"].sort(key=lambda d: d["jaar"])

with open(OUT_PATH, "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False, indent=2)

print(f"Geschreven: {OUT_PATH}")
print(f"- {len(seed['maanden'])} maand-records ({HUIDIG_JAAR})")
print(f"- vermogen snapshot: eigen vermogen check = "
      f"{spaarrekening + belegging_totaal + aandelen_payt_waarde + overwaarde_aandeel - schuld}")
print(f"- {len(seed['aandelenPayt'])} aandeelhouders")
print(f"- {len(seed['loonontwikkeling'])} loon-records")
print(f"- {len(seed['doelen'])} jaren aan doelen: "
      f"{[d['jaar'] for d in seed['doelen']]}")
